#!/usr/bin/env python3
"""
NEPSE Supertrend scan → Excel report.

TradingView Supertrend defaults: ATR period 10, multiplier 3.
A stock "falls under" Supertrend when close is below the Supertrend line
(bearish / red). A NEW SELL flip is the first close that turns the trend
from bullish to bearish.

Data sources:
  --source floorsheet  (default)  daily OHLCV from floorsheet_YYYY-MM-DD.csv
  --source nepse                  official history via nepalstock.com

Usage:
  python supertrend_scanner.py --all
  python supertrend_scanner.py --all -o supertrend_scan.xlsx
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import ssl
import sys
import time
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

ATR_PERIOD = 10
ST_MULTIPLIER = 3.0
MIN_HISTORY = 30
SLEEP_BETWEEN_SYMBOLS = 0.25
HISTORY_CALENDAR_DAYS = 500
SKIP_NAME_TOKENS = (
    "Mutual Fund",
    "Debenture",
    "Bond",
    "Promoter",
    "Preference",
    "Fund",
    "Scheme",
)
SKIP_SECTORS = {"Promoter Share", "Mutual Fund"}
PORTFOLIO_FILE = os.getenv("PORTFOLIO_FILE", "portfolio_data.json")
STOCKMAP_FILE = os.getenv("STOCKMAP_FILE", "stockmap.json")


# ── Data classes ──────────────────────────────────────────────────
@dataclass
class SupertrendRow:
    symbol: str
    name: str
    sector: str
    date: str
    open: float
    high: float
    low: float
    close: float
    volume: float
    turnover: float
    supertrend: float
    atr: float
    rsi: float
    trend: int  # 1 bullish, -1 bearish
    prev_trend: int
    days_in_trend: int
    change_1d_pct: float
    change_5d_pct: float
    vol_ma20: float
    in_portfolio: bool = False

    @property
    def below_supertrend(self) -> bool:
        return self.trend == -1

    @property
    def sell_flip(self) -> bool:
        return self.trend == -1 and self.prev_trend == 1

    @property
    def buy_flip(self) -> bool:
        return self.trend == 1 and self.prev_trend == -1

    @property
    def signal(self) -> str:
        if self.sell_flip:
            return "NEW SELL"
        if self.buy_flip:
            return "NEW BUY"
        return "BEARISH" if self.trend == -1 else "BULLISH"

    @property
    def distance_rs(self) -> float:
        return round(self.close - self.supertrend, 2)

    @property
    def distance_pct(self) -> float:
        if not self.supertrend:
            return 0.0
        return round((self.close - self.supertrend) / self.supertrend * 100.0, 2)

    @property
    def vol_ratio(self) -> float:
        if not self.vol_ma20:
            return 0.0
        return round(self.volume / self.vol_ma20, 2)


# ── Indicators ────────────────────────────────────────────────────
def _rma(series: pd.Series, length: int) -> pd.Series:
    """Wilder moving average (TradingView RMA / ATR smoothing)."""
    return series.ewm(alpha=1.0 / length, adjust=False).mean()


def compute_supertrend(
    df: pd.DataFrame,
    period: int = ATR_PERIOD,
    multiplier: float = ST_MULTIPLIER,
) -> pd.DataFrame:
    """
    TradingView ta.supertrend(factor, atrPeriod) on a daily OHLCV frame.

    Expected columns: businessDate, high, low, close. Optional: open, volume.
    Adds: atr, supertrend, st_trend (1=bullish, -1=bearish), rsi, vol_ma20.
    """
    out = df.copy().sort_values("businessDate").reset_index(drop=True)
    high = out["high"].astype(float)
    low = out["low"].astype(float)
    close = out["close"].astype(float)

    if "open" not in out.columns:
        out["open"] = close.shift(1).fillna(close)
    else:
        out["open"] = out["open"].astype(float).fillna(close.shift(1)).fillna(close)

    if "volume" not in out.columns:
        out["volume"] = 0.0
    else:
        out["volume"] = out["volume"].astype(float).fillna(0.0)

    tr = pd.concat(
        [high - low, (high - close.shift(1)).abs(), (low - close.shift(1)).abs()],
        axis=1,
    ).max(axis=1)
    atr = _rma(tr, period)
    hl2 = (high + low) / 2.0
    basic_ub = hl2 + multiplier * atr
    basic_lb = hl2 - multiplier * atr

    n = len(out)
    final_ub = np.zeros(n)
    final_lb = np.zeros(n)
    st = np.zeros(n)
    trend = np.ones(n, dtype=int)

    basic_ub_v = basic_ub.to_numpy(dtype=float)
    basic_lb_v = basic_lb.to_numpy(dtype=float)
    close_v = close.to_numpy(dtype=float)
    atr_v = atr.to_numpy(dtype=float)

    for i in range(n):
        if i == 0 or not np.isfinite(atr_v[i]):
            final_ub[i] = basic_ub_v[i] if np.isfinite(basic_ub_v[i]) else np.nan
            final_lb[i] = basic_lb_v[i] if np.isfinite(basic_lb_v[i]) else np.nan
            trend[i] = 1
            st[i] = final_lb[i]
            continue

        if basic_lb_v[i] > final_lb[i - 1] or close_v[i - 1] < final_lb[i - 1]:
            final_lb[i] = basic_lb_v[i]
        else:
            final_lb[i] = final_lb[i - 1]

        if basic_ub_v[i] < final_ub[i - 1] or close_v[i - 1] > final_ub[i - 1]:
            final_ub[i] = basic_ub_v[i]
        else:
            final_ub[i] = final_ub[i - 1]

        prev_was_upper = np.isclose(st[i - 1], final_ub[i - 1], rtol=0, atol=1e-9)
        if prev_was_upper:
            if close_v[i] <= final_ub[i]:
                trend[i] = -1
                st[i] = final_ub[i]
            else:
                trend[i] = 1
                st[i] = final_lb[i]
        else:
            if close_v[i] >= final_lb[i]:
                trend[i] = 1
                st[i] = final_lb[i]
            else:
                trend[i] = -1
                st[i] = final_ub[i]

    out["atr"] = atr
    out["supertrend"] = st
    out["st_trend"] = trend

    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = (-delta).clip(lower=0.0)
    rs = _rma(gain, 14) / _rma(loss, 14).replace(0, 1e-10)
    out["rsi"] = 100.0 - (100.0 / (1.0 + rs))
    out["vol_ma20"] = out["volume"].rolling(20, min_periods=1).mean()
    return out


def _days_in_trend(trends: np.ndarray) -> int:
    if trends.size == 0:
        return 0
    last = int(trends[-1])
    count = 0
    for v in trends[::-1]:
        if int(v) != last:
            break
        count += 1
    return count


def evaluate_latest(
    df: pd.DataFrame,
    symbol: str,
    name: str = "",
    sector: str = "",
    in_portfolio: bool = False,
    period: int = ATR_PERIOD,
    multiplier: float = ST_MULTIPLIER,
    min_history: int = MIN_HISTORY,
) -> Optional[SupertrendRow]:
    """Supertrend snapshot for the last daily bar."""
    if df is None or len(df) < min_history:
        return None
    indicated = df if "supertrend" in df.columns else compute_supertrend(df, period, multiplier)
    if indicated["supertrend"].iloc[-1] != indicated["supertrend"].iloc[-1]:
        return None

    last = indicated.iloc[-1]
    prev = indicated.iloc[-2] if len(indicated) > 1 else last
    close = float(last["close"])
    prev_close = float(prev["close"])
    close_5 = float(indicated["close"].iloc[-6]) if len(indicated) >= 6 else float(indicated["close"].iloc[0])
    bdate = last["businessDate"]
    date_str = bdate.strftime("%Y-%m-%d") if hasattr(bdate, "strftime") else str(bdate)[:10]
    turnover = float(last["turnover"]) if "turnover" in indicated.columns and pd.notna(last.get("turnover")) else 0.0

    return SupertrendRow(
        symbol=symbol,
        name=name or symbol,
        sector=sector or "Unknown",
        date=date_str,
        open=round(float(last["open"]), 2),
        high=round(float(last["high"]), 2),
        low=round(float(last["low"]), 2),
        close=round(close, 2),
        volume=round(float(last["volume"]), 0),
        turnover=round(turnover, 0),
        supertrend=round(float(last["supertrend"]), 2),
        atr=round(float(last["atr"]) if pd.notna(last["atr"]) else 0.0, 2),
        rsi=round(float(last["rsi"]) if pd.notna(last["rsi"]) else 0.0, 1),
        trend=int(last["st_trend"]),
        prev_trend=int(prev["st_trend"]),
        days_in_trend=_days_in_trend(indicated["st_trend"].to_numpy()),
        change_1d_pct=round((close - prev_close) / prev_close * 100.0, 2) if prev_close else 0.0,
        change_5d_pct=round((close - close_5) / close_5 * 100.0, 2) if close_5 else 0.0,
        vol_ma20=round(float(last["vol_ma20"]) if pd.notna(last["vol_ma20"]) else 0.0, 0),
        in_portfolio=in_portfolio,
    )


# ── Data loading ──────────────────────────────────────────────────
def load_json(path: str, default):
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError) as e:
            print(f"Warning: could not read {path}: {e}")
    return default


def load_portfolio_symbols() -> list[str]:
    data = load_json(PORTFOLIO_FILE, {})
    if isinstance(data, dict):
        return [str(s).upper() for s in data.keys()]
    return []


def load_stockmap() -> dict:
    data = load_json(STOCKMAP_FILE, {})
    return data if isinstance(data, dict) else {}


def _is_ordinary_equity_name(name: str, sector: str = "") -> bool:
    if sector in SKIP_SECTORS:
        return False
    lower = (name or "").lower()
    return not any(tok.lower() in lower for tok in SKIP_NAME_TOKENS)


def load_floorsheets(data_dir: str = ".") -> tuple[pd.DataFrame, list[str]]:
    files = sorted(glob.glob(os.path.join(data_dir, "floorsheet_20*.csv")))
    files = [f for f in files if "dividend" not in os.path.basename(f).lower()]
    if not files:
        raise FileNotFoundError("No floorsheet_YYYY-MM-DD.csv files found.")
    dfs = []
    for f in files:
        try:
            dfs.append(pd.read_csv(f, usecols=lambda c: c in {
                "stockSymbol", "contractRate", "contractQuantity",
                "contractAmount", "businessDate", "securityName",
            }))
        except Exception as e:
            print(f"  skip {f}: {e}")
    if not dfs:
        raise FileNotFoundError("Floorsheet CSVs could not be read.")
    data = pd.concat(dfs, ignore_index=True)
    data["contractAmount"] = pd.to_numeric(data["contractAmount"], errors="coerce")
    data["contractQuantity"] = pd.to_numeric(data["contractQuantity"], errors="coerce")
    data["contractRate"] = pd.to_numeric(data["contractRate"], errors="coerce")
    data["businessDate"] = pd.to_datetime(data["businessDate"])
    return data, files


def build_ohlcv(data: pd.DataFrame) -> pd.DataFrame:
    g = data.groupby(["stockSymbol", "businessDate"], sort=True)
    ohlcv = g["contractRate"].agg(open="first", high="max", low="min", close="last").reset_index()
    vol = g["contractQuantity"].sum().reset_index(name="volume")
    turnover = g["contractAmount"].sum().reset_index(name="turnover")
    ohlcv = ohlcv.merge(vol, on=["stockSymbol", "businessDate"])
    ohlcv = ohlcv.merge(turnover, on=["stockSymbol", "businessDate"])
    return ohlcv


def patch_nepse_tls(Nepse) -> None:
    def patched_request_get(self, url, include_authorization_headers=True):
        full_url = f"https://www.nepalstock.com{url}"
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json",
        }
        if include_authorization_headers:
            access_token = self.token_manager.getAccessToken()
            headers["Authorization"] = f"Salter {access_token}"
        ctx = ssl._create_unverified_context()
        req = urllib.request.Request(full_url, headers=headers)
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=30) as response:
                return json.loads(response.read().decode())
        except Exception as e:
            print(f"  NEPSE request error {url}: {e}")
            return {}

    Nepse.requestGETAPI = patched_request_get


def history_to_ohlcv(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    rename = {
        "closePrice": "close",
        "highPrice": "high",
        "lowPrice": "low",
        "openPrice": "open",
        "totalTradedQuantity": "volume",
        "totalTradedValue": "turnover",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    needed = {"businessDate", "high", "low", "close", "volume"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"history missing columns: {missing}")
    df["businessDate"] = pd.to_datetime(df["businessDate"])
    df = df.drop_duplicates("businessDate").sort_values("businessDate")
    return df


def fetch_symbol_history(n, cid: int) -> list[dict]:
    full_history: list[dict] = []
    end_date = date.today()
    start_date = end_date - timedelta(days=HISTORY_CALENDAR_DAYS)
    for page in range(0, 6):
        url = (
            f"/api/nots/market/history/security/{cid}"
            f"?&size=500&startDate={start_date}&endDate={end_date}&page={page}"
        )
        res = n.requestGETAPI(url)
        if not res or "content" not in res or not res["content"]:
            break
        full_history.extend(res["content"])
        if len(res["content"]) < 15:
            break
    return full_history


# ── Scan ──────────────────────────────────────────────────────────
def scan_ohlcv_map(
    ohlcv_by_symbol: dict[str, pd.DataFrame],
    stockmap: dict,
    portfolio: set[str],
    period: int = ATR_PERIOD,
    multiplier: float = ST_MULTIPLIER,
    min_history: int = MIN_HISTORY,
) -> tuple[list[SupertrendRow], int]:
    rows: list[SupertrendRow] = []
    skipped = 0
    for symbol, raw in ohlcv_by_symbol.items():
        meta = stockmap.get(symbol, {}) if isinstance(stockmap.get(symbol), dict) else {}
        name = str(meta.get("name") or symbol)
        sector = str(meta.get("sector") or "")
        if not _is_ordinary_equity_name(name, sector):
            skipped += 1
            continue
        if raw is None or len(raw) < min_history:
            skipped += 1
            continue
        try:
            indicated = compute_supertrend(raw, period, multiplier)
            row = evaluate_latest(
                indicated,
                symbol=symbol,
                name=name,
                sector=sector,
                in_portfolio=symbol in portfolio,
                period=period,
                multiplier=multiplier,
                min_history=min_history,
            )
        except Exception as e:
            print(f"  {symbol}: evaluate error {e}")
            skipped += 1
            continue
        if row is None:
            skipped += 1
            continue
        rows.append(row)
    return rows, skipped


def load_ohlcv_floorsheet(
    data_dir: str,
    symbols: Optional[Iterable[str]],
) -> dict[str, pd.DataFrame]:
    print("Loading floorsheet CSVs…")
    data, files = load_floorsheets(data_dir)
    print(f"  {len(files)} files, {len(data):,} trades, {data['businessDate'].nunique()} sessions")
    ohlcv = build_ohlcv(data)
    wanted = {s.upper() for s in symbols} if symbols else None
    out: dict[str, pd.DataFrame] = {}
    for symbol, grp in ohlcv.groupby("stockSymbol"):
        if wanted is not None and str(symbol).upper() not in wanted:
            continue
        out[str(symbol).upper()] = grp.drop(columns=["stockSymbol"]).copy()
    return out


def load_ohlcv_nepse(symbols: list[str]) -> dict[str, pd.DataFrame]:
    sys.setrecursionlimit(10000)
    from nepse import Nepse

    patch_nepse_tls(Nepse)
    n = Nepse()
    n.setTLSVerification(False)
    cid_map = n.getSecurityIDKeyMap() or {}
    out: dict[str, pd.DataFrame] = {}
    for i, symbol in enumerate(symbols):
        if i and i % 10 == 0:
            print(f"    Progress: {i}/{len(symbols)}")
        cid = cid_map.get(symbol)
        if not cid:
            print(f"  {symbol}: no security id")
            continue
        try:
            rows = fetch_symbol_history(n, cid)
            if rows:
                out[symbol] = history_to_ohlcv(rows)
        except Exception as e:
            print(f"  {symbol}: history error {e}")
        time.sleep(SLEEP_BETWEEN_SYMBOLS)
    return out


def resolve_nepse_universe(scan_all: bool, only: Optional[Iterable[str]]) -> list[str]:
    if only:
        return [s.strip().upper() for s in only if s.strip()]
    if not scan_all:
        return load_portfolio_symbols()
    from nepse import Nepse

    patch_nepse_tls(Nepse)
    n = Nepse()
    n.setTLSVerification(False)
    securities = n.getSecurityList() or []
    return [
        s["symbol"]
        for s in securities
        if s.get("symbol") and _is_ordinary_equity_name(
            str(s.get("securityName") or s.get("companyName") or "")
        )
    ]


# ── Excel ─────────────────────────────────────────────────────────
NAVY = "1B365D"
RED = "C0392B"
GREEN = "1E8449"
AMBER = "B9770E"
LIGHT_RED = "FADBD8"
LIGHT_GREEN = "D5F5E3"
LIGHT_AMBER = "FCF3CF"
WHITE = "FFFFFF"
GRAY = "F4F6F7"
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=18)
SUB_FONT = Font(name="Calibri", bold=True, color=NAVY, size=13)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

COLUMNS = [
    ("Symbol", 12),
    ("Company", 42),
    ("Sector", 22),
    ("Date", 13),
    ("Signal", 14),
    ("Close", 12),
    ("Supertrend", 13),
    ("Dist Rs", 12),
    ("Dist %", 11),
    ("Trend days", 12),
    ("1D %", 10),
    ("5D %", 10),
    ("RSI", 9),
    ("ATR", 10),
    ("Volume", 14),
    ("Vol vs MA20", 13),
    ("Turnover Rs", 16),
    ("Portfolio", 12),
    ("Open", 11),
    ("High", 11),
    ("Low", 11),
]


def rows_to_dataframe(rows: list[SupertrendRow]) -> pd.DataFrame:
    records = []
    for r in rows:
        records.append({
            "Symbol": r.symbol,
            "Company": r.name,
            "Sector": r.sector,
            "Date": r.date,
            "Signal": r.signal,
            "Close": r.close,
            "Supertrend": r.supertrend,
            "Dist Rs": r.distance_rs,
            "Dist %": r.distance_pct,
            "Trend days": r.days_in_trend,
            "1D %": r.change_1d_pct,
            "5D %": r.change_5d_pct,
            "RSI": r.rsi,
            "ATR": r.atr,
            "Volume": r.volume,
            "Vol vs MA20": r.vol_ratio,
            "Turnover Rs": r.turnover,
            "Portfolio": "Yes" if r.in_portfolio else "",
            "Open": r.open,
            "High": r.high,
            "Low": r.low,
        })
    df = pd.DataFrame(records, columns=[c[0] for c in COLUMNS])
    if df.empty:
        return df
    order = {"NEW SELL": 0, "BEARISH": 1, "NEW BUY": 2, "BULLISH": 3}
    df["_ord"] = df["Signal"].map(order).fillna(9)
    df = df.sort_values(["_ord", "1D %", "Dist %"], ascending=[True, True, True])
    return df.drop(columns=["_ord"]).reset_index(drop=True)


def _style_header(ws, row: int, ncol: int, fill_hex: str = NAVY):
    fill = PatternFill("solid", fgColor=fill_hex)
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[row].height = 22


def _write_df_sheet(ws, df: pd.DataFrame, table_name: str, header_hex: str = NAVY):
    ncol = len(COLUMNS)
    if df.empty:
        ws.cell(1, 1, "No stocks matched this filter.").font = Font(name="Calibri", italic=True, size=12)
        return

    for r_idx, record in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(record, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")

    _style_header(ws, 1, ncol, header_hex)

    pct_cols = {"Dist %", "1D %", "5D %"}
    num_cols = {"Close", "Supertrend", "Dist Rs", "RSI", "ATR", "Vol vs MA20", "Open", "High", "Low"}
    int_cols = {"Trend days", "Volume", "Turnover Rs"}
    col_index = {name: i + 1 for i, (name, _) in enumerate(COLUMNS)}

    for r_idx in range(2, len(df) + 2):
        signal = ws.cell(r_idx, col_index["Signal"]).value
        if signal == "NEW SELL":
            fill = PatternFill("solid", fgColor=LIGHT_AMBER)
        elif signal == "BEARISH":
            fill = PatternFill("solid", fgColor=LIGHT_RED)
        elif signal == "NEW BUY":
            fill = PatternFill("solid", fgColor="D4EFDF")
        else:
            fill = PatternFill("solid", fgColor=LIGHT_GREEN if r_idx % 2 else WHITE)
        for c_idx in range(1, ncol + 1):
            ws.cell(r_idx, c_idx).fill = fill

        for name in pct_cols:
            ws.cell(r_idx, col_index[name]).number_format = '0.00"%"'
            ws.cell(r_idx, col_index[name]).alignment = Alignment(horizontal="right")
        for name in num_cols:
            ws.cell(r_idx, col_index[name]).number_format = "#,##0.00"
        for name in int_cols:
            ws.cell(r_idx, col_index[name]).number_format = "#,##0"
        ws.cell(r_idx, col_index["Symbol"]).font = Font(name="Calibri", bold=True, size=11)
        if signal in ("NEW SELL", "BEARISH"):
            ws.cell(r_idx, col_index["Signal"]).font = Font(name="Calibri", bold=True, color=RED, size=11)
        elif signal in ("NEW BUY", "BULLISH"):
            ws.cell(r_idx, col_index["Signal"]).font = Font(name="Calibri", bold=True, color=GREEN, size=11)

    last_row = len(df) + 1
    last_col = get_column_letter(ncol)
    table = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "B2"

    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.left.text = "NEPSE Supertrend Scan"
    ws.print_title_rows = "1:1"


def _kpi_box(ws, row: int, col: int, label: str, value, fill_hex: str):
    cell_l = ws.cell(row, col, label)
    cell_v = ws.cell(row + 1, col, value)
    cell_l.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    cell_l.fill = PatternFill("solid", fgColor=fill_hex)
    cell_l.alignment = Alignment(horizontal="center")
    cell_v.font = Font(name="Calibri", bold=True, color=fill_hex, size=18)
    cell_v.fill = PatternFill("solid", fgColor="FBFCFC")
    cell_v.alignment = Alignment(horizontal="center", vertical="center")
    cell_l.border = THIN
    cell_v.border = THIN
    ws.row_dimensions[row + 1].height = 28


def write_excel(
    path: str,
    rows: list[SupertrendRow],
    as_of: str,
    scanned: int,
    skipped: int,
    source: str,
    period: int,
    multiplier: float,
    files_used: int = 0,
) -> str:
    falling = [r for r in rows if r.below_supertrend]
    sell_flips = [r for r in rows if r.sell_flip]
    buy_flips = [r for r in rows if r.buy_flip]
    bullish = [r for r in rows if r.trend == 1]
    portfolio_hits = [r for r in falling if r.in_portfolio]
    by_sector = (
        pd.Series([r.sector or "Unknown" for r in falling]).value_counts().head(12)
        if falling else pd.Series(dtype=int)
    )

    wb = Workbook()

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NEPSE Supertrend Scan Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        f"Stocks closing below Supertrend ({period}, {multiplier:g})  ·  "
        f"As of {as_of}  ·  Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, color="5D6D7E", size=11)
    ws.merge_cells("A2:G2")

    _kpi_box(ws, 4, 1, "FALLING UNDER ST", len(falling), RED)
    _kpi_box(ws, 4, 2, "NEW SELL FLIPS", len(sell_flips), AMBER)
    _kpi_box(ws, 4, 3, "NEW BUY FLIPS", len(buy_flips), GREEN)
    _kpi_box(ws, 4, 4, "STILL BULLISH", len(bullish), NAVY)
    _kpi_box(ws, 4, 5, "PORTFOLIO HITS", len(portfolio_hits), "6C3483")
    _kpi_box(ws, 4, 6, "SCANNED", scanned, "2E86AB")

    ws["A7"] = "How to read this report"
    ws["A7"].font = SUB_FONT
    notes = [
        f"Supertrend uses Wilder ATR({period}) × {multiplier:g} on daily bars (TradingView default).",
        "FALLING UNDER = last close is below the Supertrend line (bearish / red Supertrend).",
        "NEW SELL = Supertrend flipped from bullish to bearish on the latest session.",
        "NEW BUY = Supertrend flipped from bearish to bullish on the latest session.",
        "Dist % is (Close − Supertrend) / Supertrend. Negative means the stock is under the line.",
        "Trend days counts consecutive sessions in the current Supertrend direction.",
        f"Data source: {source}" + (f" ({files_used} floorsheet files)." if files_used else "."),
        f"{skipped} symbols skipped (too little history, non-equity, or missing prices).",
        "Not financial advice. Supertrend is a lagging trend filter, not a complete strategy.",
    ]
    for i, line in enumerate(notes):
        ws.cell(8 + i, 1, "• " + line).font = Font(name="Calibri", size=11, color="2C3E50")
        ws.merge_cells(start_row=8 + i, start_column=1, end_row=8 + i, end_column=7)

    start = 18
    ws.cell(start, 1, "Bearish stocks by sector").font = SUB_FONT
    ws.cell(start + 1, 1, "Sector").font = HEADER_FONT
    ws.cell(start + 1, 2, "Count").font = HEADER_FONT
    ws.cell(start + 1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(start + 1, 2).fill = PatternFill("solid", fgColor=NAVY)
    if by_sector.empty:
        ws.cell(start + 2, 1, "None")
    else:
        for i, (sector, count) in enumerate(by_sector.items(), start=start + 2):
            ws.cell(i, 1, sector).font = BODY_FONT
            ws.cell(i, 2, int(count)).font = BODY_FONT
            ws.cell(i, 2).number_format = "0"
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Falling under Supertrend by sector"
        chart.y_axis.title = None
        chart.x_axis.title = "Stocks"
        data_ref = Reference(ws, min_col=2, min_row=start + 1, max_row=start + 1 + len(by_sector))
        cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(by_sector))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 15
        chart.height = 8
        ws.add_chart(chart, "D18")

    if portfolio_hits:
        ph = start + 3 + max(len(by_sector), 1)
        ws.cell(ph, 1, "Portfolio names currently under Supertrend").font = SUB_FONT
        ws.cell(ph + 1, 1, ", ".join(sorted(r.symbol for r in portfolio_hits))).font = Font(
            name="Calibri", bold=True, color=RED, size=12
        )

    for col, width in zip("ABCDEFG", [28, 16, 16, 16, 16, 14, 18]):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.print_options.horizontalCentered = True

    sheets = [
        ("Falling Under ST", falling, RED, "FallingUnderST"),
        ("New SELL Flip", sell_flips, AMBER, "NewSellFlip"),
        ("New BUY Flip", buy_flips, GREEN, "NewBuyFlip"),
        ("All Scanned", rows, NAVY, "AllScanned"),
    ]
    for title, subset, color, table in sheets:
        ws_s = wb.create_sheet(title)
        _write_df_sheet(ws_s, rows_to_dataframe(subset), table, color)

    path = str(Path(path))
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def filter_latest_session(rows: list[SupertrendRow]) -> tuple[list[SupertrendRow], list[SupertrendRow], str]:
    """Keep bars from the latest session so stale symbols do not mix into today."""
    if not rows:
        return [], [], date.today().isoformat()
    as_of = max(r.date for r in rows)
    live = [r for r in rows if r.date == as_of]
    stale = [r for r in rows if r.date != as_of]
    return live, stale, as_of


def default_output_path(as_of: str) -> str:
    return f"supertrend_scan_{as_of}.xlsx"
    return f"supertrend_scan_{as_of}.xlsx"


def run_scan(
    scan_all: bool = True,
    symbols: Optional[list[str]] = None,
    source: str = "floorsheet",
    data_dir: str = ".",
    output: Optional[str] = None,
    period: int = ATR_PERIOD,
    multiplier: float = ST_MULTIPLIER,
) -> dict:
    stockmap = load_stockmap()
    portfolio = set(load_portfolio_symbols())
    only = [s.strip().upper() for s in symbols] if symbols else None
    files_used = 0

    if source == "nepse":
        universe = resolve_nepse_universe(scan_all=scan_all, only=only)
        if not scan_all and not only:
            print(f"Scanning portfolio ({len(universe)} symbols) via NEPSE…")
        else:
            print(f"Scanning {len(universe)} symbols via NEPSE…")
        ohlcv_map = load_ohlcv_nepse(universe)
    else:
        wanted = only
        if not scan_all and not only:
            wanted = load_portfolio_symbols()
            print(f"Scanning portfolio ({len(wanted)} symbols) from floorsheets…")
        else:
            print("Scanning all ordinary equities from floorsheets…")
        ohlcv_map = load_ohlcv_floorsheet(data_dir, wanted)
        files_used = len(glob.glob(os.path.join(data_dir, "floorsheet_20*.csv")))

    rows, skipped = scan_ohlcv_map(
        ohlcv_map, stockmap, portfolio, period=period, multiplier=multiplier
    )
    rows, stale, as_of = filter_latest_session(rows)
    skipped += len(stale)
    if stale:
        print(f"  Dropped {len(stale)} symbols with no trade on {as_of}")
    falling = [r for r in rows if r.below_supertrend]
    sell_flips = [r for r in rows if r.sell_flip]
    buy_flips = [r for r in rows if r.buy_flip]

    out_path = output or default_output_path(as_of)
    write_excel(
        out_path,
        rows,
        as_of=as_of,
        scanned=len(rows),
        skipped=skipped,
        source=source,
        period=period,
        multiplier=multiplier,
        files_used=files_used,
    )

    print(f"\nSupertrend ({period}, {multiplier:g})  as of {as_of}")
    print(f"  Scanned {len(rows)}  skipped {skipped}")
    print(f"  Falling under Supertrend: {len(falling)}")
    print(f"  New SELL flips: {len(sell_flips)}")
    print(f"  New BUY flips:  {len(buy_flips)}")
    if sell_flips:
        print("  SELL today: " + ", ".join(r.symbol for r in sell_flips))
    print(f"  Excel: {out_path}")

    return {
        "as_of": as_of,
        "output": out_path,
        "scanned": len(rows),
        "skipped": skipped,
        "falling": len(falling),
        "sell_flips": len(sell_flips),
        "buy_flips": len(buy_flips),
        "rows": rows,
    }


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="NEPSE Supertrend scan → Excel")
    parser.add_argument("--all", action="store_true", help="Scan all ordinary equities (default if no --symbols)")
    parser.add_argument("--symbols", type=str, default="", help="Comma-separated symbols")
    parser.add_argument("--source", choices=["floorsheet", "nepse"], default="floorsheet")
    parser.add_argument("--data-dir", default=".")
    parser.add_argument("-o", "--output", default="", help="Output .xlsx path")
    parser.add_argument("--period", type=int, default=ATR_PERIOD)
    parser.add_argument("--multiplier", type=float, default=ST_MULTIPLIER)
    args = parser.parse_args(argv)

    only = [s for s in args.symbols.split(",") if s.strip()] or None
    scan_all = args.all or only is None
    payload = run_scan(
        scan_all=scan_all if only is None else False,
        symbols=only,
        source=args.source,
        data_dir=args.data_dir,
        output=args.output or None,
        period=args.period,
        multiplier=args.multiplier,
    )
    return 0 if payload.get("scanned", 0) else 1


if __name__ == "__main__":
    raise SystemExit(main())
