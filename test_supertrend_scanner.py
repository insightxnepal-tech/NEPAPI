#!/usr/bin/env python3
"""Unit tests for Supertrend math, scan filters, and Excel export (no NEPSE network)."""

import os
import tempfile
import unittest

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import supertrend_scanner as st


def make_ohlcv(n=80, start=100.0, trend=0.4, crash=False, recover=False):
    """Synthetic daily bars. Uptrend by default; optional crash / recovery."""
    dates = pd.bdate_range("2025-01-01", periods=n)
    t = np.linspace(0, 1, n)
    close = start * (1 + trend * t)
    if crash:
        close[-8:] = close[-9] * np.linspace(0.92, 0.72, 8)
    if recover:
        close[-8:] = close[-9] * np.linspace(1.02, 1.22, 8)
    close = np.maximum(close, 1.0)
    open_ = np.r_[close[0], close[:-1]]
    high = np.maximum(open_, close) * 1.004
    low = np.minimum(open_, close) * 0.996
    volume = np.full(n, 12_000.0)
    volume[-1] = 20_000.0
    return pd.DataFrame({
        "businessDate": dates,
        "open": open_,
        "high": high,
        "low": low,
        "close": close,
        "volume": volume,
        "turnover": volume * close,
    })


class SupertrendMathTests(unittest.TestCase):
    def test_uptrend_closes_above_supertrend(self):
        df = st.compute_supertrend(make_ohlcv(trend=0.5))
        self.assertEqual(int(df["st_trend"].iloc[-1]), 1)
        self.assertGreater(df["close"].iloc[-1], df["supertrend"].iloc[-1])

    def test_crash_flips_to_bearish(self):
        df = st.compute_supertrend(make_ohlcv(trend=0.5, crash=True))
        self.assertEqual(int(df["st_trend"].iloc[-1]), -1)
        self.assertLess(df["close"].iloc[-1], df["supertrend"].iloc[-1])

    def test_sell_flip_detected(self):
        raw = make_ohlcv(trend=0.5)
        prev = float(raw["close"].iloc[-2])
        raw.loc[raw.index[-1], ["open", "high", "low", "close"]] = [
            prev * 0.95, prev * 0.96, prev * 0.70, prev * 0.72,
        ]
        row = st.evaluate_latest(raw, "TEST")
        self.assertIsNotNone(row)
        self.assertTrue(row.below_supertrend)
        self.assertEqual(row.signal, "NEW SELL")
        self.assertTrue(row.sell_flip)

    def test_buy_flip_after_recovery(self):
        raw = make_ohlcv(trend=0.5, crash=True)
        self.assertEqual(int(st.compute_supertrend(raw)["st_trend"].iloc[-2]), -1)
        prev = float(raw["close"].iloc[-2])
        raw.loc[raw.index[-1], ["open", "high", "low", "close"]] = [
            prev * 1.05, prev * 1.40, prev * 1.04, prev * 1.38,
        ]
        row = st.evaluate_latest(raw, "TEST")
        self.assertIsNotNone(row)
        self.assertEqual(row.trend, 1)
        self.assertTrue(row.buy_flip)
        self.assertEqual(row.signal, "NEW BUY")

    def test_short_history_skipped(self):
        df = make_ohlcv(n=10)
        self.assertIsNone(st.evaluate_latest(df, "TINY"))

    def test_days_in_trend_counts_tail(self):
        trends = np.array([1, 1, -1, -1, -1, -1])
        self.assertEqual(st._days_in_trend(trends), 4)

    def test_supertrend_is_one_of_the_bands(self):
        df = st.compute_supertrend(make_ohlcv())
        last = df.iloc[-1]
        self.assertTrue(np.isfinite(last["supertrend"]))
        self.assertGreater(last["atr"], 0)

    def test_distance_pct_negative_when_below(self):
        row = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "TEST")
        self.assertLess(row.distance_pct, 0)
        self.assertLess(row.distance_rs, 0)


class ScanAndExcelTests(unittest.TestCase):
    def test_scan_splits_falling_under(self):
        bull = make_ohlcv(trend=0.5)
        bear = make_ohlcv(trend=0.5, crash=True)
        rows, skipped = st.scan_ohlcv_map(
            {"BULL": bull, "BEAR": bear},
            stockmap={"BULL": {"name": "Bull Co", "sector": "Banks"},
                      "BEAR": {"name": "Bear Co", "sector": "Hydro Power"}},
            portfolio={"BEAR"},
        )
        self.assertEqual(skipped, 0)
        by_sym = {r.symbol: r for r in rows}
        self.assertEqual(by_sym["BULL"].signal, "BULLISH")
        self.assertTrue(by_sym["BEAR"].below_supertrend)
        self.assertTrue(by_sym["BEAR"].in_portfolio)

    def test_skips_growth_fund_from_stockmap(self):
        rows, skipped = st.scan_ohlcv_map(
            {"NICGF2": make_ohlcv()},
            stockmap={"NICGF2": {"name": "NIC ASIA Growth Fund-2", "sector": "Commercial Banks"}},
            portfolio=set(),
        )
        self.assertEqual(rows, [])
        self.assertEqual(skipped, 1)

    def test_skips_mutual_fund_sector(self):
        rows, skipped = st.scan_ohlcv_map(
            {"SAGF": make_ohlcv()},
            stockmap={"SAGF": {"name": "Sanima Growth Fund", "sector": "Mutual Fund"}},
            portfolio=set(),
        )
        self.assertEqual(rows, [])
        self.assertEqual(skipped, 1)

    def test_excel_workbook_sheets_and_falling_rows(self):
        bull = st.evaluate_latest(make_ohlcv(trend=0.5), "BULL", name="Bull Co", sector="Banks")
        bear = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "BEAR", name="Bear Co", sector="Hydro Power")
        self.assertIsNotNone(bull)
        self.assertIsNotNone(bear)
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "scan.xlsx")
            st.write_excel(
                path,
                [bull, bear],
                as_of=bear.date,
                scanned=2,
                skipped=0,
                source="floorsheet",
                period=10,
                multiplier=3.0,
                files_used=3,
            )
            self.assertTrue(os.path.exists(path))
            self.assertGreater(os.path.getsize(path), 4000)
            wb = load_workbook(path)
            self.assertEqual(
                wb.sheetnames,
                ["Summary", "Best Setups", "Falling Under ST", "New SELL Flip", "New BUY Flip", "All Scanned"],
            )
            falling = wb["Falling Under ST"]
            symbols = [falling.cell(r, 1).value for r in range(2, falling.max_row + 1)]
            self.assertIn("BEAR", symbols)
            self.assertNotIn("BULL", symbols)
            all_sheet = wb["All Scanned"]
            self.assertEqual(all_sheet.max_row, 3)  # header + 2
            self.assertEqual(wb["Summary"]["A1"].value, "NEPSE Supertrend Scan Report")

    def test_filter_latest_session_drops_stale(self):
        live = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "LIVE")
        stale = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "OLD")
        live.date = "2026-09-02"
        stale.date = "2026-08-01"
        kept, dropped, as_of = st.filter_latest_session([live, stale])
        self.assertEqual([r.symbol for r in kept], ["LIVE"])
        self.assertEqual([r.symbol for r in dropped], ["OLD"])
        self.assertEqual(as_of, live.date)

    def test_rows_to_dataframe_puts_sell_first(self):
        bull = st.evaluate_latest(make_ohlcv(trend=0.5), "BULL")
        bear = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "BEAR")
        df = st.rows_to_dataframe([bull, bear])
        self.assertEqual(df.iloc[0]["Symbol"], "BEAR")
        self.assertIn(df.iloc[0]["Signal"], ("NEW SELL", "BEARISH"))
        self.assertEqual(df.iloc[1]["Signal"], "BULLISH")

    def test_classify_setup_best_buy_needs_breadth_and_pullback(self):
        row = st.evaluate_latest(make_ohlcv(trend=0.5), "BULL")
        row.ema21 = row.close - 1
        row.vol_ma20 = row.volume / 1.5
        row.rsi = 52.0
        # distance_pct is a property from close/supertrend — pin close near ST
        row.supertrend = row.close / 1.008
        row.days_in_trend = 5
        row.open = row.close - 1
        weak = st.classify_setup(row, breadth_pct=30.0)
        self.assertEqual(weak.label, "WATCH BUY")
        strong = st.classify_setup(row, breadth_pct=50.0)
        self.assertEqual(strong.label, "BEST BUY")

    def test_classify_setup_best_sell_fresh_breakdown(self):
        row = st.evaluate_latest(make_ohlcv(trend=0.5, crash=True), "BEAR")
        row.ema21 = row.close + 10
        row.vol_ma20 = max(row.volume / 1.5, 1)
        row.rsi = 35.0
        row.days_in_trend = 1
        row.prev_trend = 1
        row.trend = -1
        row.open = row.close + 5
        grade = st.classify_setup(row, breadth_pct=37.0)
        self.assertEqual(grade.label, "BEST SELL")
        self.assertEqual(grade.kind, "sell")


if __name__ == "__main__":
    unittest.main()
